import math
import openpyxl
import cmath

K2u=openpyxl.load_workbook(r"C:\Users\nicki\Desktop\диплом\Шины7ТПС\etap-TDULF-sub7-%VUF2-19.5.0E.xlsx")
ListK2u = K2u.active

Va = openpyxl.load_workbook(r"C:\Users\nicki\Desktop\диплом\Шины7ТПС\etap-TDULF-sub7-Voltage(PhaseA)-19.5.0E.xlsx")
ListVa = Va.active

Vb = openpyxl.load_workbook(r"C:\Users\nicki\Desktop\диплом\Шины7ТПС\etap-TDULF-sub7-Voltage(PhaseB)-19.5.0E.xlsx")
ListVb = Vb.active

Vc = openpyxl.load_workbook(r"C:\Users\nicki\Desktop\диплом\Шины7ТПС\etap-TDULF-sub7-Voltage(PhaseC)-19.5.0E.xlsx")
ListVc = Vc.active

anglea = openpyxl.load_workbook(r"C:\Users\nicki\Desktop\диплом\Шины7ТПС\etap-TDULF-sub7-VoltageAngle(PhaseA)-19.5.0E.xlsx")
Listanglea = anglea.active

angleb = openpyxl.load_workbook(r"C:\Users\nicki\Desktop\диплом\Шины7ТПС\etap-TDULF-sub7-VoltageAngle(PhaseB)-19.5.0E.xlsx")
Listangleb = angleb.active

anglec = openpyxl.load_workbook(r"C:\Users\nicki\Desktop\диплом\Шины7ТПС\etap-TDULF-sub7-VoltageAngle(PhaseC)-19.5.0E.xlsx")
Listanglec = anglec.active

from openpyxl import Workbook

wb = Workbook()
wl = wb.active
wl.cell(row=1, column=1).value="Beta"
wl.cell(row=1, column=2).value="K2U"

a=complex(-0.5 , 0.866)
for i in range(2,600):
    ra=ListVa.cell(row=i, column=2).value*math.cos(Listanglea.cell(row=i, column=2).value*3.14/180)
    ima=ListVa.cell(row=i, column=2).value*math.sin(Listanglea.cell(row=i, column=2).value*3.14/180)
    Va=complex(ra,ima)

    rb=ListVb.cell(row=i, column=2).value*math.cos(Listangleb.cell(row=i, column=2).value*3.14/180)
    imb=ListVb.cell(row=i, column=2).value*math.sin(Listangleb.cell(row=i, column=2).value*3.14/180)
    Vb=complex(rb,imb)

    rc=ListVc.cell(row=i, column=2).value*math.cos(Listanglec.cell(row=i, column=2).value*3.14/180)
    imc=ListVc.cell(row=i, column=2).value*math.sin(Listanglec.cell(row=i, column=2).value*3.14/180)
    Vc=complex(rc,imc)

    V1=(Va+a*Vb+a*a*Vc)/3
    V2=(Va+a*a*Vb+a*Vc)/3
    Beta=(cmath.phase(V1)-cmath.phase(V2))*180/3.14
    wl.cell(row=i, column=1).value = Beta
    wl.cell(row=i, column=2).value = ListK2u.cell(row=i, column=2).value
    
wb.save(r'C:\Users\nicki\Desktop\диплом\Шины7ТПС\ТПС7.xlsx')

