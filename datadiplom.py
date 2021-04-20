import openpyxl
from openpyxl import Workbook
wb = Workbook()
wl = wb.active
nps=1   #номер ПС
klf=2   #количество фидеров
kf=klf+1
for i in range(1,kf):
    s=r'C:\Users\nicki\Desktop\etap\данные\ФКС'+str(i)+' ТП'+str(nps)+'-Amps(FromA).xlsx'
    rb = openpyxl.load_workbook(s)
    rl = rb.active
    if rl.cell(2,2).value==0:
        s=r'C:\Users\nicki\Desktop\etap\данные\ФКС'+str(i)+' ТП'+str(nps)+'-Amps(FromB).xlsx'
        rb = openpyxl.load_workbook(s)
        rl = rb.active
        fase='BC'
    else:
        s=r'C:\Users\nicki\Desktop\etap\данные\ФКС'+str(i)+' ТП'+str(nps)+'-Amps(FromB).xlsx'
        rb = openpyxl.load_workbook(s)
        rl = rb.active
        if rl.cell(2,2).value==0:
            fase='AC'
            s=r'C:\Users\nicki\Desktop\etap\данные\ФКС'+str(i)+' ТП'+str(nps)+'-Amps(FromC).xlsx'
            rb = openpyxl.load_workbook(s)
            rl = rb.active
        else:
            fase='AB'
    for j in range(2,602):
        a=rl.cell(row=j, column=2).value
        wl.cell(row=j, column=1).value = a*25/1000*0.9
        wl.cell(row=j, column=2).value = a*25/1000*0.4358899
        a=rl.cell(row=j, column=1).value
        wl.cell(row=j, column=10).value = int(a[11])
        wl.cell(row=j, column=11).value = int(a[13:15])
        wl.cell(row=j, column=12).value = int(a[16:18])
        wl.cell(row=j, column=13).value = a[:11]
    rb = openpyxl.load_workbook(r'C:\Users\nicki\Desktop\etap\éaáTÑ¡¿Ñ pú½á U2\TractionTimeDomain\TDULF\Lump1.xlsx')
    rl = rb.active
    for j in range(1,14):
        a=rl.cell(row=1, column=j).value
        wl.cell(row=1, column=j).value = a
    wb.save(r'C:\Users\nicki\Desktop\etap\данные\TDULF\f'+str(i)+'s'+str(nps)+'fase'+fase+'.xlsx')
