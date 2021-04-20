import openpyxl

rb = openpyxl.load_workbook(r'C:\Users\nicki\Desktop\экзампл\лист1.xlsx')
rl = rb.active

from openpyxl import Workbook

wb = Workbook()
wl = wb.active

for i in range(2,22):
    a = rl.cell(row=i, column=2).value
    wl.cell(row=i, column=1).value = a*2
    a = rl.cell(row=i, column=1).value
    print(a)
    wl.cell(row=i, column=2).value = a[3:]

wb.save(r'C:\Users\nicki\Desktop\экзампл\попа.xlsx')
